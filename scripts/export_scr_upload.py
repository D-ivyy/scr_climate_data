#!/usr/bin/env python3
"""Export InfraSure dev DB assets into the SCR climate-data import template.

The vendor-facing file intentionally carries a generated SCR asset name. The
private manifest is the join-back layer to InfraSure UUIDs, workspaces, and
portfolios.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

try:
    import psycopg2
    import psycopg2.extras
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit("Missing dependency: psycopg2. Run with the platform .venv.") from exc

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - dependency guard
    raise SystemExit("Missing dependency: openpyxl. Run with the platform .venv.") from exc


SCRIPT_PATH = Path(__file__).resolve()
SCR_ROOT = SCRIPT_PATH.parents[1]
LAB_ROOT = SCR_ROOT.parent
DEFAULT_TEMPLATE = SCR_ROOT / "docs" / "Climate_Metrics_Import_tempate.xlsx"
DEFAULT_ENV_FILE = SCR_ROOT / ".env"
DEFAULT_OUTPUT_ROOT = SCR_ROOT / "runs"

ASSETS_SHEET = "Assets"
FIRST_DATA_ROW = 10
LAST_TEMPLATE_ROW = 1010

FORMULA_COLUMNS = {"D", "E", "F", "G", "K"}

XLSX_COLUMN_BY_KEY = {
    "asset_name": "B",
    "asset_type": "C",
    "ticcs_class": "D",
    "ticcs_class_name": "E",
    "nace_code": "F",
    "sector": "G",
    "operating_entity_name": "H",
    "detention_rate": "I",
    "country": "J",
    "country_code": "K",
    "address": "L",
    "latitude": "M",
    "longitude": "N",
    "revenues": "O",
    "asset_value": "P",
    "year": "Q",
    "scope_1": "R",
    "scope_2": "S",
    "decarbonisation_scope_1": "T",
    "decarbonisation_scope_2": "U",
    "flood_blue_green_infrastructure": "V",
    "flood_drainage_system_upgrades": "W",
    "flood_elevation": "X",
    "flood_erosion_control": "Y",
    "flood_barriers": "Z",
    "flood_pumps": "AA",
    "flood_reservoir_capacity_expansion": "AB",
    "flood_sea_walls": "AC",
    "flood_spillway_enhancement": "AD",
    "flood_structural_improvements": "AE",
    "flood_undergrounding": "AF",
    "storm_vegetation_management": "AG",
    "storm_structural_improvements": "AH",
    "storm_undergrounding": "AI",
    "storm_wind_breaks_structural": "AJ",
    "storm_wind_breaks_terrain_landscaping": "AK",
    "storm_wind_breaks_vegetation": "AL",
    "heat_blue_green_infrastructure": "AM",
    "heat_cooling_systems_mechanical": "AN",
    "heat_cooling_systems_natural_evaporative": "AO",
    "heat_reflective_coatings": "AP",
    "heat_resistant_construction_materials": "AQ",
    "heat_insulation": "AR",
    "heat_optimise_operational_practices": "AS",
    "heat_shading_greening": "AT",
    "heat_shading_shade_structures": "AU",
    "heat_undergrounding": "AV",
    "wildfire_defensible_space_management": "AW",
    "wildfire_erosion_sediment_control": "AX",
    "wildfire_fire_retardants": "AY",
    "wildfire_fire_suppression_systems": "AZ",
    "wildfire_firebreaks": "BA",
    "wildfire_fireproof_building_materials": "BB",
    "wildfire_fuel_reduction_zones": "BC",
    "wildfire_prescribed_burns": "BD",
    "wildfire_structural_hardening": "BE",
    "wildfire_undergrounding": "BF",
    "wildfire_vegetation_management_landscaping": "BG",
}

CSV_COLUMNS = [
    "Asset Name",
    "Asset type",
    "TICCS Class",
    "TICCS Class name",
    "NACE code",
    "Sector",
    "Operating entity name",
    "Detention rate",
    "Country",
    "Country Code (ISO3)",
    "Address",
    "Latitude",
    "Longitude",
    "Revenues",
    "Asset Value",
    "Year",
    "Scope 1",
    "Scope 2",
    "Decarbonisation Scope 1",
    "Decarbonisation Scope 2",
    "Flood: Blue-green infrastructure",
    "Flood: Drainage system upgrades",
    "Flood: Elevation",
    "Flood: Erosion control",
    "Flood: Flood barriers",
    "Flood: Pumps",
    "Flood: Reservoir capacity expansion",
    "Flood: Sea walls",
    "Flood: Spillway enhancement",
    "Flood: Structural improvements",
    "Flood: Undergrounding",
    "Storm: Natural infrastructure - vegetation management",
    "Storm: Structural improvements",
    "Storm: Undergrounding",
    "Storm: Wind breaks - Structural",
    "Storm: Wind breaks - terrain and landscaping",
    "Storm: Wind breaks - vegetation",
    "Heat: Blue-green infrastructure",
    "Heat: Cooling systems - mechanical",
    "Heat: Cooling systems - natural and evaporative",
    "Heat: Heat reflective coatings",
    "Heat: Heat-resistant construction materials",
    "Heat: Insulation",
    "Heat: Optimise operational practices",
    "Heat: Shading - greening",
    "Heat: Shading - shade structures",
    "Heat: Undergrounding",
    "Wildfire: Defensible space management",
    "Wildfire: Erosion and sediment control",
    "Wildfire: Fire retardants",
    "Wildfire: Fire suppression systems",
    "Wildfire: Firebreaks",
    "Wildfire: Fireproof building materials",
    "Wildfire: Fuel reduction zones",
    "Wildfire: Prescribed burns",
    "Wildfire: Structural hardening",
    "Wildfire: Undergrouding",
    "Wildfire: Vegetation management and landscaping",
]

CSV_KEY_BY_HEADER = {
    "Asset Name": "asset_name",
    "Asset type": "asset_type",
    "Operating entity name": "operating_entity_name",
    "Detention rate": "detention_rate",
    "Country": "country",
    "Address": "address",
    "Latitude": "latitude",
    "Longitude": "longitude",
    "Revenues": "revenues",
    "Asset Value": "asset_value",
    "Year": "year",
    "Scope 1": "scope_1",
    "Scope 2": "scope_2",
    "Decarbonisation Scope 1": "decarbonisation_scope_1",
    "Decarbonisation Scope 2": "decarbonisation_scope_2",
    "Flood: Blue-green infrastructure": "flood_blue_green_infrastructure",
    "Flood: Drainage system upgrades": "flood_drainage_system_upgrades",
    "Flood: Elevation": "flood_elevation",
    "Flood: Erosion control": "flood_erosion_control",
    "Flood: Flood barriers": "flood_barriers",
    "Flood: Pumps": "flood_pumps",
    "Flood: Reservoir capacity expansion": "flood_reservoir_capacity_expansion",
    "Flood: Sea walls": "flood_sea_walls",
    "Flood: Spillway enhancement": "flood_spillway_enhancement",
    "Flood: Structural improvements": "flood_structural_improvements",
    "Flood: Undergrounding": "flood_undergrounding",
    "Storm: Natural infrastructure - vegetation management": "storm_vegetation_management",
    "Storm: Structural improvements": "storm_structural_improvements",
    "Storm: Undergrounding": "storm_undergrounding",
    "Storm: Wind breaks - Structural": "storm_wind_breaks_structural",
    "Storm: Wind breaks - terrain and landscaping": "storm_wind_breaks_terrain_landscaping",
    "Storm: Wind breaks - vegetation": "storm_wind_breaks_vegetation",
    "Heat: Blue-green infrastructure": "heat_blue_green_infrastructure",
    "Heat: Cooling systems - mechanical": "heat_cooling_systems_mechanical",
    "Heat: Cooling systems - natural and evaporative": "heat_cooling_systems_natural_evaporative",
    "Heat: Heat reflective coatings": "heat_reflective_coatings",
    "Heat: Heat-resistant construction materials": "heat_resistant_construction_materials",
    "Heat: Insulation": "heat_insulation",
    "Heat: Optimise operational practices": "heat_optimise_operational_practices",
    "Heat: Shading - greening": "heat_shading_greening",
    "Heat: Shading - shade structures": "heat_shading_shade_structures",
    "Heat: Undergrounding": "heat_undergrounding",
    "Wildfire: Defensible space management": "wildfire_defensible_space_management",
    "Wildfire: Erosion and sediment control": "wildfire_erosion_sediment_control",
    "Wildfire: Fire retardants": "wildfire_fire_retardants",
    "Wildfire: Fire suppression systems": "wildfire_fire_suppression_systems",
    "Wildfire: Firebreaks": "wildfire_firebreaks",
    "Wildfire: Fireproof building materials": "wildfire_fireproof_building_materials",
    "Wildfire: Fuel reduction zones": "wildfire_fuel_reduction_zones",
    "Wildfire: Prescribed burns": "wildfire_prescribed_burns",
    "Wildfire: Structural hardening": "wildfire_structural_hardening",
    "Wildfire: Undergrouding": "wildfire_undergrounding",
    "Wildfire: Vegetation management and landscaping": "wildfire_vegetation_management_landscaping",
}

EIA_FUEL_TO_ASSET_TYPE = {
    "SUN": "Photovoltaic power generation facility",
    "WND": "On-shore wind power generation facility",
    "WAT": "Hydroelectric dam",
    "GEO": "Geothermal power plant",
    "NUC": "Nuclear power plant",
    "NG": "Gas power Plant",
    "BFG": "Gas power Plant",
    "OG": "Gas power Plant",
    "PG": "Gas power Plant",
    "BIT": "Coal power Plant",
    "SUB": "Coal power Plant",
    "LIG": "Coal power Plant",
    "WC": "Coal power Plant",
    "RC": "Coal power Plant",
    "DFO": "Other power plant",
    "RFO": "Other power plant",
    "JF": "Other power plant",
    "KER": "Other power plant",
    "WO": "Other power plant",
    "MWH": "Battery storage unit",
    "OBG": "Biomass power plant",
    "OBL": "Biomass power plant",
    "OBS": "Biomass power plant",
    "AB": "Biomass power plant",
    "WDL": "Biomass power plant",
    "WDS": "Biomass power plant",
    "LFG": "Biomass power plant",
    "MSW": "Waste to power facility",
}

CSV_KEY_ALIASES = {
    "scr_asset_name": "scr_asset_name",
    "asset_name": "scr_asset_name",
    "plant_uuid": "plant_uuid",
    "plant_id": "plant_uuid",
    "portfolio_asset_id": "portfolio_asset_id",
    "eia_plant_id": "eia_plant_id",
    "asset_value": "asset_value",
    "asset_value_usd": "asset_value",
    "valuation_usd": "asset_value",
    "revenues": "revenues",
    "revenue": "revenues",
    "revenue_usd": "revenues",
    "annual_revenue_usd": "revenues",
    "year": "year",
    "valuation_year": "year",
    "detention_rate": "detention_rate",
    "operating_entity_name": "operating_entity_name",
    "operator": "operating_entity_name",
    "address": "address",
    "asset_type": "asset_type",
    "scope_1": "scope_1",
    "scope_1_tco2e": "scope_1",
    "scope_2": "scope_2",
    "scope_2_tco2e": "scope_2",
}


@dataclass
class DbAsset:
    workspace_id: str | None
    workspace_slug: str | None
    portfolio_id: str | None
    portfolio_name: str | None
    portfolio_kind: str | None
    portfolio_asset_id: str | None
    workspace_asset_state: str | None
    plant_uuid: str
    eia_plant_id: int | None
    plant_slug: str
    source_type: str
    plant_name: str | None
    operating_entity_name: str | None
    state: str | None
    county: str | None
    city: str | None
    postal_code: str | None
    address_raw: str | None
    latitude: Decimal | None
    longitude: Decimal | None
    primary_fuel: str | None
    fuel_types: list[str]
    technologies: list[str]
    total_capacity_mw: Decimal | None
    latest_capex_per_kw: Decimal | None


@dataclass
class ExportRow:
    values: dict[str, Any]
    manifest: dict[str, Any]
    errors: list[str]
    warnings: list[str]

    @property
    def accepted(self) -> bool:
        return not self.errors


def read_env_file(path: Path | None) -> dict[str, str]:
    if not path or not path.exists():
        return {}
    out: dict[str, str] = {}
    for raw_line in path.read_text().splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        out[key.strip()] = value.strip().strip('"').strip("'")
    return out


def resolve_database_url(args: argparse.Namespace) -> str:
    if args.database_url:
        return args.database_url
    env_file = Path(args.env_file).expanduser() if args.env_file else DEFAULT_ENV_FILE
    env_values = read_env_file(env_file)
    if args.target == "prod":
        url = os.environ.get("DATABASE_URL") or env_values.get("DATABASE_URL")
        if not url:
            raise SystemExit("DATABASE_URL not found. Pass --database-url or --env-file.")
        return url
    url = os.environ.get("DATABASE_URL_DEV") or env_values.get("DATABASE_URL_DEV")
    if not url:
        raise SystemExit("DATABASE_URL_DEV not found. Pass --database-url or --env-file.")
    return url


def safe_db_host(url: str) -> str:
    parsed = urlparse(url)
    return parsed.hostname or ""


def normalize_key(value: str) -> str:
    key = value.strip().lower()
    key = re.sub(r"[^a-z0-9]+", "_", key)
    return key.strip("_")


def as_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def decimal_to_number(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def decimal_to_text(value: Decimal | None) -> str:
    if value is None:
        return ""
    return format(value.normalize(), "f")


def array_from_db(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v) for v in value if v is not None]
    if isinstance(value, tuple):
        return [str(v) for v in value if v is not None]
    return [str(value)]


def slugify_short(text: str | None, fallback: str = "asset") -> str:
    base = normalize_key(text or fallback).replace("_", "-")
    base = re.sub(r"-+", "-", base).strip("-")
    return base or fallback


def scr_asset_name(asset: DbAsset, max_length: int = 120) -> str:
    suffix = f"is_p_{asset.plant_uuid.replace('-', '')}"
    slug = slugify_short(asset.plant_slug or asset.plant_name)
    name = f"{slug}__{suffix}"
    if len(name) <= max_length:
        return name
    return suffix


def load_template_maps(template_path: Path) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    wb = load_workbook(template_path, data_only=True, read_only=True)
    asset_types: dict[str, dict[str, str]] = {}
    ws = wb["Asset types"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        asset_type = row[0]
        if not asset_type:
            continue
        asset_types[str(asset_type)] = {
            "nace_code": row[3] or "",
            "sector": row[4] or "",
            "ticcs_class": row[8] or "",
            "ticcs_class_name": row[9] or "",
        }
    country_codes: dict[str, str] = {}
    ws = wb["Country code"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            country_codes[str(row[0])] = str(row[1])
    return asset_types, country_codes


def load_sidecar(path: Path | None) -> dict[str, dict[str, str]]:
    if not path:
        return {}
    if not path.exists():
        raise SystemExit(f"Asset sidecar not found: {path}")

    records: dict[str, dict[str, str]] = {}
    with path.open(newline="") as fh:
        reader = csv.DictReader(fh)
        for raw in reader:
            normalized: dict[str, str] = {}
            for key, value in raw.items():
                if key is None:
                    continue
                canonical = CSV_KEY_ALIASES.get(normalize_key(key), normalize_key(key))
                normalized[canonical] = (value or "").strip()

            keys = [
                normalized.get("plant_uuid"),
                normalized.get("scr_asset_name"),
                normalized.get("portfolio_asset_id"),
                normalized.get("eia_plant_id"),
            ]
            for key in keys:
                if key:
                    records[key] = normalized
    return records


def sidecar_for_asset(sidecar: dict[str, dict[str, str]], asset: DbAsset, generated_name: str) -> dict[str, str]:
    keys = [
        asset.plant_uuid,
        generated_name,
        asset.portfolio_asset_id,
        str(asset.eia_plant_id) if asset.eia_plant_id is not None else None,
    ]
    for key in keys:
        if key and key in sidecar:
            return sidecar[key]
    return {}


def resolve_workspace_id(cur: Any, args: argparse.Namespace) -> str | None:
    if args.workspace_id:
        return args.workspace_id
    if not args.workspace_slug:
        return None
    cur.execute(
        "SELECT id::text FROM workspace WHERE slug=%s AND deleted_at IS NULL",
        (args.workspace_slug,),
    )
    row = cur.fetchone()
    if not row:
        raise SystemExit(f"Workspace slug not found: {args.workspace_slug}")
    return row["id"]


def resolve_portfolio_id(cur: Any, args: argparse.Namespace, workspace_id: str | None) -> str | None:
    if args.portfolio_id:
        return args.portfolio_id
    if not args.portfolio_name:
        return None
    if not workspace_id:
        raise SystemExit("--portfolio-name requires --workspace-slug or --workspace-id")
    cur.execute(
        """
        SELECT id::text
          FROM portfolio
         WHERE workspace_id=%s::uuid
           AND name=%s
           AND deleted_at IS NULL
         ORDER BY created_at DESC
        """,
        (workspace_id, args.portfolio_name),
    )
    rows = cur.fetchall()
    if not rows:
        raise SystemExit(f"Portfolio not found in workspace: {args.portfolio_name}")
    if len(rows) > 1:
        raise SystemExit(f"Portfolio name is not unique in workspace: {args.portfolio_name}")
    return rows[0]["id"]


def selection_query(args: argparse.Namespace, workspace_id: str | None, portfolio_id: str | None) -> tuple[str, list[Any]]:
    base_select = """
      SELECT
        w.id::text AS workspace_id,
        w.slug AS workspace_slug,
        pf.id::text AS portfolio_id,
        pf.name AS portfolio_name,
        pf.kind AS portfolio_kind,
        pa.id::text AS portfolio_asset_id,
        wa.state AS workspace_asset_state,
        p.id::text AS plant_uuid,
        p.eia_plant_id,
        p.slug AS plant_slug,
        p.source_type,
        COALESCE(NULLIF(p.data->>'plant_name',''), NULLIF(p.data->>'name',''), pi.plant_name) AS plant_name,
        COALESCE(NULLIF(p.data->>'utility_eia_name',''), pi.utility_eia_name, pi.canonical_owner, pi.gem_parent_name) AS operating_entity_name,
        COALESCE(NULLIF(p.data->>'state',''), pi.state) AS state,
        COALESCE(NULLIF(p.data->>'county',''), pi.county) AS county,
        COALESCE(NULLIF(p.data->>'city',''), NULLIF(p.data->>'municipality','')) AS city,
        COALESCE(NULLIF(p.data->>'zip_code',''), NULLIF(p.data->>'postal_code','')) AS postal_code,
        COALESCE(NULLIF(p.data->>'address',''), NULLIF(p.data->>'street_address',''), NULLIF(p.data->>'location','')) AS address_raw,
        COALESCE(NULLIF(p.data->>'lat',''), NULLIF(p.data->>'latitude',''), pi.lat::text) AS latitude,
        COALESCE(NULLIF(p.data->>'lon',''), NULLIF(p.data->>'longitude',''), pi.lon::text) AS longitude,
        COALESCE(NULLIF(p.data->>'primary_fuel',''), pi.primary_fuel) AS primary_fuel,
        COALESCE(pi.fuel_types, ARRAY[]::text[]) AS fuel_types,
        COALESCE(pi.technologies, ARRAY[]::text[]) AS technologies,
        COALESCE(NULLIF(p.data->>'total_capacity_mw','')::numeric, pi.total_capacity_mw::numeric) AS total_capacity_mw,
        fp.latest_capex_per_kw
      FROM plants p
      LEFT JOIN plant_index pi ON pi.plant_id = p.eia_plant_id
      LEFT JOIN financial_plant fp ON fp.plant_id = p.id
    """

    if args.all_reference:
        sql = base_select + """
          LEFT JOIN workspace w ON false
          LEFT JOIN workspace_asset wa ON false
          LEFT JOIN portfolio pf ON false
          LEFT JOIN portfolio_asset pa ON false
         WHERE p.workspace_id IS NULL
        """
        params: list[Any] = []
    elif args.plant_uuid:
        sql = base_select + """
          LEFT JOIN workspace w ON w.id = p.workspace_id
          LEFT JOIN workspace_asset wa ON wa.workspace_id = w.id AND wa.plant_id = p.id
          LEFT JOIN portfolio_asset pa ON pa.reference_plant_id = p.id AND pa.deleted_at IS NULL
          LEFT JOIN portfolio pf ON pf.id = pa.portfolio_id AND pf.deleted_at IS NULL
         WHERE p.id = ANY(%s::uuid[])
        """
        params = [args.plant_uuid]
    elif portfolio_id:
        sql = base_select + """
          JOIN portfolio_asset pa ON pa.reference_plant_id = p.id AND pa.deleted_at IS NULL
          JOIN portfolio pf ON pf.id = pa.portfolio_id AND pf.deleted_at IS NULL
          JOIN workspace w ON w.id = pa.workspace_id AND w.deleted_at IS NULL
          LEFT JOIN workspace_asset wa ON wa.workspace_id = pa.workspace_id AND wa.plant_id = p.id
         WHERE pa.portfolio_id = %s::uuid
        """
        params = [portfolio_id]
    else:
        if not workspace_id:
            raise SystemExit(
                "Select assets with --portfolio-id/--portfolio-name, --workspace-slug, "
                "--plant-uuid, or --all-reference."
            )
        sql = base_select + """
          JOIN workspace_asset wa ON wa.plant_id = p.id
          JOIN workspace w ON w.id = wa.workspace_id AND w.deleted_at IS NULL
          LEFT JOIN portfolio_asset pa ON pa.workspace_id = wa.workspace_id
                                     AND pa.reference_plant_id = p.id
                                     AND pa.deleted_at IS NULL
                                     AND pa.is_baseline IS TRUE
          LEFT JOIN portfolio pf ON pf.id = pa.portfolio_id AND pf.deleted_at IS NULL
         WHERE wa.workspace_id = %s::uuid
        """
        params = [workspace_id]
        if args.asset_state != "all":
            sql += " AND wa.state = %s"
            params.append(args.asset_state)

    sql += " ORDER BY COALESCE(w.slug, ''), COALESCE(pf.name, ''), p.slug"
    if args.limit:
        sql += " LIMIT %s"
        params.append(args.limit)
    return sql, params


def fetch_assets(conn: Any, args: argparse.Namespace) -> tuple[list[DbAsset], str | None, str | None]:
    with conn.cursor() as cur:
        workspace_id = resolve_workspace_id(cur, args)
        portfolio_id = resolve_portfolio_id(cur, args, workspace_id)
        sql, params = selection_query(args, workspace_id, portfolio_id)
        cur.execute(sql, params)
        rows = cur.fetchall()

    assets: list[DbAsset] = []
    for row in rows:
        assets.append(
            DbAsset(
                workspace_id=row["workspace_id"],
                workspace_slug=row["workspace_slug"],
                portfolio_id=row["portfolio_id"],
                portfolio_name=row["portfolio_name"],
                portfolio_kind=row["portfolio_kind"],
                portfolio_asset_id=row["portfolio_asset_id"],
                workspace_asset_state=row["workspace_asset_state"],
                plant_uuid=row["plant_uuid"],
                eia_plant_id=row["eia_plant_id"],
                plant_slug=row["plant_slug"],
                source_type=row["source_type"],
                plant_name=row["plant_name"],
                operating_entity_name=row["operating_entity_name"],
                state=row["state"],
                county=row["county"],
                city=row["city"],
                postal_code=row["postal_code"],
                address_raw=row["address_raw"],
                latitude=as_decimal(row["latitude"]),
                longitude=as_decimal(row["longitude"]),
                primary_fuel=row["primary_fuel"],
                fuel_types=array_from_db(row["fuel_types"]),
                technologies=array_from_db(row["technologies"]),
                total_capacity_mw=as_decimal(row["total_capacity_mw"]),
                latest_capex_per_kw=as_decimal(row["latest_capex_per_kw"]),
            )
        )
    return assets, workspace_id, portfolio_id


def infer_asset_type(asset: DbAsset) -> str:
    codes = [asset.primary_fuel or ""] + asset.fuel_types
    code_set = {str(c).upper() for c in codes if c}
    tech_text = " ".join(asset.technologies).lower()
    name_text = (asset.plant_name or asset.plant_slug or "").lower()

    if "MWH" in code_set or "battery" in tech_text or "battery" in name_text:
        return "Battery storage unit"
    if "SUN" in code_set:
        if "thermal solar" in tech_text or "solar thermal" in tech_text:
            return "Thermal solar power facility"
        return "Photovoltaic power generation facility"
    if "WND" in code_set:
        if "offshore" in tech_text or "off-shore" in name_text or "offshore" in name_text:
            return "Off-shore wind power generation facility"
        return "On-shore wind power generation facility"
    if "WAT" in code_set:
        if "pumped" in tech_text or "pumped" in name_text:
            return "Pumped hydroelectric storage facility"
        return "Hydroelectric dam"
    for code in codes:
        mapped = EIA_FUEL_TO_ASSET_TYPE.get(str(code).upper())
        if mapped:
            return mapped
    return "Other power plant"


def compose_address(asset: DbAsset, sidecar_value: str | None = None) -> tuple[str, str]:
    if sidecar_value:
        return sidecar_value, "sidecar"
    if asset.address_raw:
        parts = [asset.address_raw, asset.city, asset.state, asset.postal_code, "USA"]
        return ", ".join(str(p).strip() for p in parts if p), "plants.data.address"
    if asset.city and asset.state:
        parts = [asset.city, asset.state, asset.postal_code, "USA"]
        return ", ".join(str(p).strip() for p in parts if p), "city_state_fallback"
    if asset.county and asset.state:
        county = asset.county if "county" in asset.county.lower() else f"{asset.county} County"
        return f"{county}, {asset.state}, USA", "county_state_fallback"
    if asset.state:
        return f"{asset.state}, USA", "state_fallback"
    return "", "missing"


def sidecar_decimal(record: dict[str, str], key: str) -> Decimal | None:
    return as_decimal(record.get(key))


def build_export_rows(
    assets: list[DbAsset],
    args: argparse.Namespace,
    sidecar: dict[str, dict[str, str]],
    asset_type_map: dict[str, dict[str, str]],
    country_codes: dict[str, str],
    run_id: str,
    generated_at: str,
) -> list[ExportRow]:
    rows: list[ExportRow] = []
    for asset in assets:
        generated_name = scr_asset_name(asset, args.max_asset_name_length)
        side = sidecar_for_asset(sidecar, asset, generated_name)
        errors: list[str] = []
        warnings: list[str] = []

        asset_type = side.get("asset_type") or infer_asset_type(asset)
        if asset_type not in asset_type_map:
            errors.append(f"asset_type_not_in_template:{asset_type}")

        address, address_source = compose_address(asset, side.get("address"))
        if not address:
            warnings.append("missing_address")

        country = side.get("country") or "United States"
        country_code = country_codes.get(country, "")
        if not country_code:
            warnings.append(f"country_code_not_found:{country}")

        detention_rate = sidecar_decimal(side, "detention_rate")
        if detention_rate is None:
            detention_rate = as_decimal(args.detention_rate)

        revenues = sidecar_decimal(side, "revenues")
        revenue_source = "sidecar" if revenues is not None else "missing"
        if revenues is None:
            warnings.append("missing_revenues_template_marks_mandatory")

        asset_value = sidecar_decimal(side, "asset_value")
        asset_value_source = "sidecar" if asset_value is not None else "missing"
        if asset_value is None and args.asset_value_source == "capex_proxy":
            if asset.total_capacity_mw is not None and asset.latest_capex_per_kw is not None:
                asset_value = asset.total_capacity_mw * asset.latest_capex_per_kw * Decimal("1000")
                asset_value_source = "capex_proxy_latest_capex_per_kw_times_mw_times_1000"
                warnings.append("asset_value_uses_capex_proxy_not_true_valuation")
            else:
                warnings.append("missing_asset_value_capex_proxy_unavailable")
        elif asset_value is None:
            warnings.append("missing_asset_value")

        year = side.get("year") or args.valuation_year

        if not generated_name:
            errors.append("missing_asset_name")
        if asset.latitude is None:
            errors.append("missing_latitude")
        if asset.longitude is None:
            errors.append("missing_longitude")
        if detention_rate is None:
            warnings.append("missing_detention_rate")
        if args.strict and revenues is None:
            errors.append("missing_revenues_strict")
        if args.require_asset_value and asset_value is None:
            errors.append("missing_asset_value")
        elif args.strict and asset_value is None:
            errors.append("missing_asset_value_strict")

        operating_entity_name = ""
        if args.include_operating_entity:
            operating_entity_name = side.get("operating_entity_name") or asset.operating_entity_name or ""

        values: dict[str, Any] = {
            "asset_name": generated_name,
            "asset_type": asset_type,
            "operating_entity_name": operating_entity_name,
            "detention_rate": decimal_to_number(detention_rate),
            "country": country,
            "address": address,
            "latitude": decimal_to_number(asset.latitude),
            "longitude": decimal_to_number(asset.longitude),
            "revenues": decimal_to_number(revenues),
            "asset_value": decimal_to_number(asset_value),
            "year": int(year) if str(year).isdigit() else year,
        }

        for key in XLSX_SIDELOAD_KEYS:
            if key in side and side[key] != "":
                values[key] = side[key]

        formula_values = asset_type_map.get(asset_type, {})
        values["ticcs_class"] = formula_values.get("ticcs_class", "")
        values["ticcs_class_name"] = formula_values.get("ticcs_class_name", "")
        values["nace_code"] = formula_values.get("nace_code", "")
        values["sector"] = formula_values.get("sector", "")
        values["country_code"] = country_code

        manifest = {
            "scr_run_id": run_id,
            "generated_at": generated_at,
            "scr_asset_name": generated_name,
            "scr_grain": "plant",
            "plant_uuid": asset.plant_uuid,
            "generator_uuid": "",
            "plant_slug": asset.plant_slug,
            "eia_plant_id": asset.eia_plant_id or "",
            "eia_generator_code": "",
            "workspace_id": asset.workspace_id or "",
            "workspace_slug": asset.workspace_slug or "",
            "portfolio_id": asset.portfolio_id or "",
            "portfolio_name": asset.portfolio_name or "",
            "portfolio_asset_id": asset.portfolio_asset_id or "",
            "workspace_asset_state": asset.workspace_asset_state or "",
            "source_type": asset.source_type,
            "plant_name": asset.plant_name or "",
            "asset_type": asset_type,
            "address": address,
            "address_source": address_source,
            "latitude": decimal_to_text(asset.latitude),
            "longitude": decimal_to_text(asset.longitude),
            "detention_rate": decimal_to_text(detention_rate),
            "revenues": decimal_to_text(revenues),
            "revenue_source": revenue_source,
            "asset_value": decimal_to_text(asset_value),
            "asset_value_source": asset_value_source,
            "valuation_year": year,
            "primary_fuel": asset.primary_fuel or "",
            "fuel_types": "|".join(asset.fuel_types),
            "technologies": "|".join(asset.technologies),
            "total_capacity_mw": decimal_to_text(asset.total_capacity_mw),
            "latest_capex_per_kw": decimal_to_text(asset.latest_capex_per_kw),
            "validation_status": "accepted" if not errors else "rejected",
            "validation_errors": "|".join(errors),
            "validation_warnings": "|".join(warnings),
        }
        rows.append(ExportRow(values=values, manifest=manifest, errors=errors, warnings=warnings))
    return rows


XLSX_SIDELOAD_KEYS = {
    "scope_1",
    "scope_2",
    "decarbonisation_scope_1",
    "decarbonisation_scope_2",
    "flood_blue_green_infrastructure",
    "flood_drainage_system_upgrades",
    "flood_elevation",
    "flood_erosion_control",
    "flood_barriers",
    "flood_pumps",
    "flood_reservoir_capacity_expansion",
    "flood_sea_walls",
    "flood_spillway_enhancement",
    "flood_structural_improvements",
    "flood_undergrounding",
    "storm_vegetation_management",
    "storm_structural_improvements",
    "storm_undergrounding",
    "storm_wind_breaks_structural",
    "storm_wind_breaks_terrain_landscaping",
    "storm_wind_breaks_vegetation",
    "heat_blue_green_infrastructure",
    "heat_cooling_systems_mechanical",
    "heat_cooling_systems_natural_evaporative",
    "heat_reflective_coatings",
    "heat_resistant_construction_materials",
    "heat_insulation",
    "heat_optimise_operational_practices",
    "heat_shading_greening",
    "heat_shading_shade_structures",
    "heat_undergrounding",
    "wildfire_defensible_space_management",
    "wildfire_erosion_sediment_control",
    "wildfire_fire_retardants",
    "wildfire_fire_suppression_systems",
    "wildfire_firebreaks",
    "wildfire_fireproof_building_materials",
    "wildfire_fuel_reduction_zones",
    "wildfire_prescribed_burns",
    "wildfire_structural_hardening",
    "wildfire_undergrounding",
    "wildfire_vegetation_management_landscaping",
}


def output_dir_for_run(args: argparse.Namespace, run_id: str) -> Path:
    if args.output_dir:
        return Path(args.output_dir).expanduser()
    return DEFAULT_OUTPUT_ROOT / run_id


def write_csv(path: Path, rows: list[ExportRow]) -> None:
    accepted = [row for row in rows if row.accepted]
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in accepted:
            out: dict[str, Any] = {}
            for header in CSV_COLUMNS:
                if header == "TICCS Class":
                    out[header] = row.values.get("ticcs_class", "")
                elif header == "TICCS Class name":
                    out[header] = row.values.get("ticcs_class_name", "")
                elif header == "NACE code":
                    out[header] = row.values.get("nace_code", "")
                elif header == "Sector":
                    out[header] = row.values.get("sector", "")
                elif header == "Country Code (ISO3)":
                    out[header] = row.values.get("country_code", "")
                else:
                    key = CSV_KEY_BY_HEADER.get(header)
                    out[header] = row.values.get(key, "") if key else ""
            writer.writerow(out)


def copy_row_style(ws: Any, source_row: int, target_row: int) -> None:
    for col in range(2, 60):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)


def write_xlsx(
    template_path: Path,
    path: Path,
    rows: list[ExportRow],
    *,
    include_support_sheets: bool,
) -> None:
    accepted = [row for row in rows if row.accepted]
    if len(accepted) > (LAST_TEMPLATE_ROW - FIRST_DATA_ROW + 1):
        raise SystemExit(
            f"Template supports {LAST_TEMPLATE_ROW - FIRST_DATA_ROW + 1} rows; "
            f"selection has {len(accepted)} accepted rows. Split the run."
        )

    wb = load_workbook(template_path, data_only=False, read_only=False)
    ws = wb[ASSETS_SHEET]

    for row_idx in range(FIRST_DATA_ROW, LAST_TEMPLATE_ROW + 1):
        for key, col in XLSX_COLUMN_BY_KEY.items():
            ws[f"{col}{row_idx}"] = None

    for offset, row in enumerate(accepted):
        row_idx = FIRST_DATA_ROW + offset
        if row_idx > FIRST_DATA_ROW:
            copy_row_style(ws, FIRST_DATA_ROW, row_idx)
        for key, col in XLSX_COLUMN_BY_KEY.items():
            ws[f"{col}{row_idx}"] = row.values.get(key)

    try:
        wb.calculation.fullCalcOnLoad = True
    except AttributeError:
        pass
    if not include_support_sheets:
        for sheet_name in list(wb.sheetnames):
            if sheet_name != ASSETS_SHEET:
                del wb[sheet_name]
    wb.save(path)


def write_manifest(path: Path, rows: list[ExportRow]) -> None:
    fieldnames = list(rows[0].manifest.keys()) if rows else [
        "scr_run_id",
        "generated_at",
        "scr_asset_name",
        "validation_status",
    ]
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.manifest)


def write_rejects(path: Path, rows: list[ExportRow]) -> None:
    rejects = [row for row in rows if not row.accepted]
    fieldnames = list(rows[0].manifest.keys()) if rows else [
        "scr_run_id",
        "generated_at",
        "scr_asset_name",
        "validation_status",
    ]
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rejects:
            writer.writerow(row.manifest)


def write_run_manifest(
    path: Path,
    args: argparse.Namespace,
    run_id: str,
    generated_at: str,
    db_host: str,
    workspace_id: str | None,
    portfolio_id: str | None,
    rows: list[ExportRow],
) -> None:
    accepted = [row for row in rows if row.accepted]
    rejected = [row for row in rows if not row.accepted]
    warned = [row for row in rows if row.warnings]
    payload = {
        "scr_run_id": run_id,
        "generated_at": generated_at,
        "target": args.target,
        "db_host": db_host,
        "workspace_id": workspace_id,
        "workspace_slug": args.workspace_slug,
        "portfolio_id": portfolio_id,
        "portfolio_name": args.portfolio_name,
        "all_reference": args.all_reference,
        "plant_uuid": args.plant_uuid,
        "asset_value_source": args.asset_value_source,
        "include_operating_entity": args.include_operating_entity,
        "strict": args.strict,
        "template": str(Path(args.template).expanduser()),
        "selected_rows": len(rows),
        "accepted_rows": len(accepted),
        "rejected_rows": len(rejected),
        "rows_with_warnings": len(warned),
        "warning_counts": count_messages(row.warnings for row in rows),
        "error_counts": count_messages(row.errors for row in rows),
    }
    path.write_text(json.dumps(payload, indent=2) + "\n")


def count_messages(groups: Any) -> dict[str, int]:
    counts: dict[str, int] = {}
    for messages in groups:
        for message in messages:
            counts[message] = counts.get(message, 0) + 1
    return dict(sorted(counts.items()))


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export InfraSure dev DB assets to SCR climate-data CSV/XLSX."
    )
    parser.add_argument("--target", choices=["dev", "prod"], default="dev")
    parser.add_argument("--database-url", help="Postgres URL. Defaults to DATABASE_URL_DEV from env/.env file.")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV_FILE), help="Env file with DATABASE_URL_DEV.")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="SCR Excel import template.")
    parser.add_argument("--output-dir", help="Output run directory. Defaults to scr_climate_data/runs/<run_id>.")
    parser.add_argument("--run-id", help="Stable run identifier. Defaults to timestamp.")

    scope = parser.add_mutually_exclusive_group()
    scope.add_argument("--all-reference", action="store_true", help="Export all reference plants.")
    scope.add_argument("--plant-uuid", action="append", help="Export a specific plants.id UUID; repeatable.")
    parser.add_argument("--workspace-id", help="Workspace UUID.")
    parser.add_argument("--workspace-slug", help="Workspace slug.")
    parser.add_argument("--portfolio-id", help="Portfolio UUID.")
    parser.add_argument("--portfolio-name", help="Portfolio name within workspace.")
    parser.add_argument("--asset-state", choices=["portfolio", "prospect", "all"], default="portfolio")
    parser.add_argument("--limit", type=int, help="Limit selected rows for testing.")

    parser.add_argument("--asset-values", help="CSV sidecar keyed by plant_uuid, scr_asset_name, portfolio_asset_id, or eia_plant_id.")
    parser.add_argument(
        "--asset-value-source",
        choices=["sidecar", "capex_proxy"],
        default="sidecar",
        help="sidecar leaves missing values blank; capex_proxy derives latest_capex_per_kw * MW * 1000 when sidecar is absent.",
    )
    parser.add_argument("--valuation-year", default=str(datetime.now().year))
    parser.add_argument("--detention-rate", default="", help="Optional template detention-rate percentage.")
    parser.add_argument(
        "--include-operating-entity",
        action="store_true",
        help="Populate optional Operating entity name from sidecar/DB. Default omits it from upload files.",
    )
    parser.add_argument(
        "--require-asset-value",
        action="store_true",
        help="Reject rows missing Asset Value. Recommended for actual SCR uploads.",
    )
    parser.add_argument("--max-asset-name-length", type=int, default=40)
    parser.add_argument("--strict", action="store_true", help="Reject rows missing template financial fields.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    template_path = Path(args.template).expanduser()
    if not template_path.exists():
        raise SystemExit(f"Template not found: {template_path}")

    run_id = args.run_id or datetime.now(timezone.utc).strftime("scr_%Y%m%dT%H%M%SZ")
    generated_at = datetime.now(timezone.utc).isoformat()
    out_dir = output_dir_for_run(args, run_id)
    out_dir.mkdir(parents=True, exist_ok=True)

    db_url = resolve_database_url(args)
    asset_type_map, country_codes = load_template_maps(template_path)
    sidecar = load_sidecar(Path(args.asset_values).expanduser() if args.asset_values else None)

    conn = psycopg2.connect(db_url, cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = True
    try:
        assets, workspace_id, portfolio_id = fetch_assets(conn, args)
    finally:
        conn.close()

    rows = build_export_rows(
        assets=assets,
        args=args,
        sidecar=sidecar,
        asset_type_map=asset_type_map,
        country_codes=country_codes,
        run_id=run_id,
        generated_at=generated_at,
    )

    csv_path = out_dir / "scr_upload.csv"
    xlsx_path = out_dir / "scr_upload.xlsx"
    template_xlsx_path = out_dir / "scr_upload_template.xlsx"
    manifest_path = out_dir / "scr_manifest.csv"
    rejects_path = out_dir / "scr_rejects.csv"
    run_manifest_path = out_dir / "run_manifest.json"

    write_csv(csv_path, rows)
    write_xlsx(template_path, xlsx_path, rows, include_support_sheets=False)
    write_xlsx(template_path, template_xlsx_path, rows, include_support_sheets=True)
    write_manifest(manifest_path, rows)
    write_rejects(rejects_path, rows)
    write_run_manifest(
        run_manifest_path,
        args=args,
        run_id=run_id,
        generated_at=generated_at,
        db_host=safe_db_host(db_url),
        workspace_id=workspace_id,
        portfolio_id=portfolio_id,
        rows=rows,
    )

    accepted = sum(1 for row in rows if row.accepted)
    rejected = len(rows) - accepted
    warned = sum(1 for row in rows if row.warnings)
    print(f"SCR run: {run_id}")
    print(f"Selected rows: {len(rows)}")
    print(f"Accepted rows: {accepted}")
    print(f"Rejected rows: {rejected}")
    print(f"Rows with warnings: {warned}")
    print(f"Wrote: {csv_path}")
    print(f"Wrote: {xlsx_path}")
    print(f"Wrote: {template_xlsx_path}")
    print(f"Wrote: {manifest_path}")
    print(f"Wrote: {rejects_path}")
    print(f"Wrote: {run_manifest_path}")
    if rejected:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
